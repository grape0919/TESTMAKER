import sys
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QMainWindow, QApplication
from view.main import Ui_MainWindow
from openpyxl import load_workbook, Workbook, styles
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.page import PrintPageSetup, PrintOptions

import pandas as pd
from pathlib import Path
import random
class WindowClass(QMainWindow, Ui_MainWindow) :

    def __init__(self) :
        super(WindowClass, self).__init__()
        super(Ui_MainWindow, self).__init__()
        self.setupUi(self)
        self.buttonGenTest.clicked.connect(self.clickGenTestButton)
        self.buttonOpenWordsFile.clicked.connect(self.clickOpenFileButton)

    def clickOpenFileButton(self):
        print("openFile")
        fname = QFileDialog.getOpenFileName(self, 'Open Words File', 'Desktop',
                                            "excel (*.xlsx *.xls)")
        if fname[0]:
            self._filePath = fname[0]
            self.editFilePath.setText(fname[0])
            
            #파일 열기
            try:
                load_wb = load_workbook(fname[0], data_only=True)
            except:
                QMessageBox.about(self, "Warning", self._filePath + "\n단어장 파일을 여는데 실패하였습니다.")
                return

            try:
                load_ws = load_wb['단어']
            except KeyError:
                QMessageBox.about(self, "Warning", "'단어' 시트가 존재하지 않습니다.")
                return

            self.all_values = {}
            for row in load_ws.rows:
                row_value = []
                for index in range(3):
                    row_value.append(row[index].value)
                try:
                    self.all_values[int(row_value[0])].append([row_value[1], row_value[2]])
                except ValueError:
                    print("데이터 형식이 맞지 않습니다. [회차, 영어, 한글]")
                except KeyError:
                    self.all_values[int(row_value[0])] = []
                    self.all_values[int(row_value[0])].append([row_value[1], row_value[2]])
                
            print(self.all_values)
            # get_cells = load_ws['A1':'C1']
            # for row in get_cells:
            #         for cell in row:
            #             print(cell.value)

            #TODO FPS 도 label로 표시

            load_wb.close()
        else:
            QMessageBox.about(self, "Warning", "파일을 선택하지 않았습니다.")
            self.editFilePath.setText("")
    
    def clickGenTestButton(self):
        print("genTest")
        if ((not self.RB_English.isChecked() and not self.RB_Korean.isChecked() and not self.RB_EK.isChecked()) 
            or not self.editFilePath.text() or not self.editFrom.text() or not self.editTo.text() or not self.editNumOfWord.text()):

            QMessageBox.about(self, "Warning", "모든 설정을 해야합니다.")
        
        else:
            
            self.outputFilePath = str(Path(self._filePath).parent)+'\\'+Path(self._filePath).name.replace(' ', '_').split('.')[0] + self.editFrom.text() + '-' + self.editTo.text() +'_Test.xlsx'
            print('outputFilePath : ', self.outputFilePath)
            tempList = []
            for i in range(int(self.editFrom.text()), int(self.editTo.text())+1):
                for j in range(len(self.all_values[i])) :
                    tempList.append(self.all_values[i][j])
            # tempList =  self.all_values[int(self.editFrom.text()):int(self.editTo.text())]
            testWordList = random.sample(tempList, int(self.editNumOfWord.text()))

            #print("testWordList: " , testWordList)

            font_styles = styles.Font(bold=True)
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            

            # writer = pd.ExcelWriter(self._filePath, engine='openpyxl')

            # writer.book = load_workbook(self._filePath)
            
            # # if truncate_sheet and sheet_name in writer.book.sheetnames:
        
            # #     idx = writer.book.sheetnames.index(sheet_name)
                
            # #     writer.book.remove(writer.book.worksheets[idx])
                
            # #     writer.book.create_sheet(sheet_name, idx)

            # # copy existing sheets
            # writer.sheets = {ws.title:ws for ws in writer.book.worksheets}


            write_wb = Workbook() 
 
            # 이름이 있는 시트를 생성
            # write_ws = writer.book.create_sheet('시험지')
            write_ws = write_wb.create_sheet('시험지')
            del_ws = write_wb['Sheet']
            write_wb.remove(del_ws)
            
            write_ws['B3'] = '이름 : '
            # write_ws['B3'].font = write_ws['B3'].font.copy(bold=True)
            write_ws['B4'] = '범위 : '
            # write_ws['B4'].font = write_ws['B4'].font.copy(font_styles)
            write_ws['C4'] = self.editFrom.text() + " - " + self.editTo.text()
            # write_ws['C4'].font = write_ws['C4'].font.copy(font_styles)

            write_ws.column_dimensions['A'].width = 5
            write_ws.column_dimensions['B'].width = 5 
            write_ws.column_dimensions['C'].width = 30
            write_ws.column_dimensions['D'].width = 30
            write_ws.column_dimensions['E'].width = 5
            write_ws.column_dimensions['F'].width = 30
            write_ws.column_dimensions['G'].width = 30
            write_ws.column_dimensions['H'].width = 5
        

            numOfWord = int(self.editNumOfWord.text())
            halfNumOfWord = int((numOfWord+1)/2)

            
            write_ws.row_dimensions[5].height = 30

            for row in write_ws.iter_rows(min_row=1, max_col=7, max_row=5):
                for cell in row:
                    cell.font = font_styles
            
            for row in write_ws.iter_rows(min_row=6, min_col=2, max_col=7, max_row=halfNumOfWord + 5):
                for cell in row:
                    cell.border = thin_border



            if self.RB_English.isChecked() :
                for i in range(halfNumOfWord):
                    write_ws.row_dimensions[i+6].height = 30
                    write_ws.cell(i+6, 2, i+1)
                    write_ws.cell(i+6, 3, testWordList[i][0])
                    write_ws.cell(i+6, 5, (i+halfNumOfWord+1))
                    if i+halfNumOfWord < numOfWord :
                        write_ws.cell(i+6, 6, testWordList[i+halfNumOfWord][0])
                    
            elif self.RB_Korean.isChecked() :
                for i in range(halfNumOfWord):
                    write_ws.row_dimensions[i+6].height = 30
                    write_ws.cell(i+6, 2, i+1)
                    write_ws.cell(i+6, 3, testWordList[i][1])
                    if i+halfNumOfWord < numOfWord :
                        write_ws.cell(i+6, 6, testWordList[i+halfNumOfWord][1])


            elif self.RB_EK.isChecked() :
                twThree = int((numOfWord / 30) * 23)
                seven = numOfWord - twThree
                print("twThree : ", twThree)
                print("seven : ", seven)
                for i in range(twThree, twThree + seven) :
                    testWordList[i].reverse()

                random.shuffle(testWordList)

                for i in range(halfNumOfWord):
                    write_ws.row_dimensions[i+6].height = 30
                    write_ws.cell(i+6, 2, i+1)
                    write_ws.cell(i+6, 3, testWordList[i][0])
                    write_ws.cell(i+6, 5, (i+halfNumOfWord+1))
                    if i+halfNumOfWord < numOfWord :
                        write_ws.cell(i+6, 6, testWordList[i+halfNumOfWord][0])

                
            write_ws.row_dimensions[halfNumOfWord+7].height = 30
            write_ws.row_dimensions[halfNumOfWord+7].height = 30

            write_ws.page_setup = PrintPageSetup(worksheet=write_ws, scale=50)  
            write_ws.print_options = PrintOptions(gridLinesSet=True)

            # 답안지 생성
            # 이름이 있는 시트를 생성
            # write_ws = writer.book.create_sheet('시험지')
            write_ws = write_wb.create_sheet('답안지')
            
            write_ws['B3'] = '이름 : '
            # write_ws['B3'].font = write_ws['B3'].font.copy(bold=True)
            write_ws['B4'] = '범위 : '
            # write_ws['B4'].font = write_ws['B4'].font.copy(font_styles)
            write_ws['C4'] = self.editFrom.text() + " - " + self.editTo.text()
            # write_ws['C4'].font = write_ws['C4'].font.copy(font_styles)

            write_ws.column_dimensions['A'].width = 5
            write_ws.column_dimensions['B'].width = 5 
            write_ws.column_dimensions['C'].width = 30
            write_ws.column_dimensions['D'].width = 30
            write_ws.column_dimensions['E'].width = 5
            write_ws.column_dimensions['F'].width = 30
            write_ws.column_dimensions['G'].width = 30
            write_ws.column_dimensions['H'].width = 5
        

            numOfWord = int(self.editNumOfWord.text())
            halfNumOfWord = int((numOfWord+1)/2)

            
            write_ws.row_dimensions[5].height = 30

            for row in write_ws.iter_rows(min_row=1, max_col=7, max_row=5):
                for cell in row:
                    cell.font = font_styles
            
            for row in write_ws.iter_rows(min_row=6, min_col=2, max_col=7, max_row=halfNumOfWord + 5):
                for cell in row:
                    cell.border = thin_border


            for i in range(halfNumOfWord):
                write_ws.row_dimensions[i+6].height = 30
                write_ws.cell(i+6, 2, i+1)
                write_ws.cell(i+6, 3, testWordList[i][0])
                write_ws.cell(i+6, 4, testWordList[i][1])
                write_ws.cell(i+6, 5, (i+halfNumOfWord+1))
                if i+halfNumOfWord < numOfWord :
                    write_ws.cell(i+6, 6, testWordList[i+halfNumOfWord][0])
                    write_ws.cell(i+6, 7, testWordList[i+halfNumOfWord][1])
                
            write_ws.row_dimensions[halfNumOfWord+7].height = 30
            write_ws.row_dimensions[halfNumOfWord+7].height = 30

            write_ws.page_setup = PrintPageSetup(worksheet=write_ws, scale=50)  
            write_ws.print_options = PrintOptions(gridLinesSet=True)
            

            try:
                write_wb.save(self.outputFilePath)
                QMessageBox.about(self, "시험지 생성 성공", self.outputFilePath+"\n시험지가 생성되었습니다.")

            except PermissionError:
                QMessageBox.about(self, "Warning", self.outputFilePath+"\n파일을 사용중이거나, 권한이 없습니다.")



            

if __name__ == "__main__" :
    #QApplication : 프로그램을 실행시켜주는 클래스
    app = QApplication(sys.argv) 

    #WindowClass의 인스턴스 생성
    myWindow = WindowClass()
    #프로그램 화면을 보여주는 코드
    myWindow.show()

    #프로그램을 이벤트루프로 진입시키는(프로그램을 작동시키는) 코드
    app.exec_()
